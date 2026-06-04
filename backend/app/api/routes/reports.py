# ============================================================
# app/api/routes/reports.py
# GET /api/reports/high-risk-summary    → PDF
# GET /api/reports/interventions        → Excel
# GET /api/reports/analytics            → PDF
# ============================================================
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from io import BytesIO
from datetime import datetime

from app.db.database import get_db
from app.services.auth_service import require_role
from app.models.user import User, UserRole
from app.models.student import Student
from app.models.intervention import Intervention, InterventionStatus
from app.core.config import settings

router = APIRouter(prefix="/reports", tags=["Reports"])


def _make_pdf_high_risk(students) -> BytesIO:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("EduGuard AI — High Risk Student Summary", styles["Title"]))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}", styles["Normal"]))
    elements.append(Spacer(1, 20))

    data = [["Student ID", "Name", "Class", "Semester", "Risk Score", "Attendance", "Factors"]]
    for s in students:
        data.append([
            s.id, s.name, s.class_name, s.semester,
            f"{s.risk_score:.0f}",
            f"{s.attendance:.0f}%",
            ", ".join(s.risk_factors),
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    return buf


def _make_excel_interventions(interventions) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Interventions"

    headers = ["Student ID", "Student Name", "Type", "Assigned By",
               "Date Assigned", "Status", "Outcome", "Notes"]

    blue_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = blue_fill
        cell.alignment = Alignment(horizontal="center")

    for row, iv in enumerate(interventions, 2):
        ws.cell(row=row, column=1, value=iv.student_id)
        ws.cell(row=row, column=2, value=iv.student.name if iv.student else "")
        ws.cell(row=row, column=3, value=iv.type.value)
        ws.cell(row=row, column=4, value=iv.assigned_by)
        ws.cell(row=row, column=5, value=iv.date_assigned.strftime("%Y-%m-%d") if iv.date_assigned else "")
        ws.cell(row=row, column=6, value=iv.status.value)
        ws.cell(row=row, column=7, value=iv.outcome.value)
        ws.cell(row=row, column=8, value=iv.notes or "")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.get("/high-risk-summary", summary="Download High Risk Summary PDF")
def high_risk_pdf(
    db: Session = Depends(get_db),
    _: User = Depends(require_role(UserRole.admin, UserRole.counselor)),
):
    students = db.query(Student).filter(
        Student.risk_score >= settings.HIGH_RISK_THRESHOLD
    ).order_by(Student.risk_score.desc()).all()

    pdf = _make_pdf_high_risk(students)
    return StreamingResponse(
        pdf,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=high_risk_summary.pdf"}
    )


@router.get("/interventions", summary="Download Intervention Progress Excel")
def interventions_excel(
    db: Session = Depends(get_db),
    _: User = Depends(require_role(UserRole.admin, UserRole.counselor)),
):
    from sqlalchemy.orm import joinedload
    ivs = db.query(Intervention).options(joinedload(Intervention.student)).all()
    buf = _make_excel_interventions(ivs)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=interventions_report.xlsx"}
    )


@router.get("/analytics", summary="Download Risk Analytics PDF")
def analytics_pdf(
    db: Session = Depends(get_db),
    _: User = Depends(require_role(UserRole.admin, UserRole.counselor)),
):
    students = db.query(Student).filter(
        Student.risk_score >= settings.HIGH_RISK_THRESHOLD
    ).order_by(Student.risk_score.desc()).all()

    pdf = _make_pdf_high_risk(students)
    return StreamingResponse(
        pdf,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=risk_analytics_report.pdf"}
    )


@router.get("/monthly-trend", summary="Download Monthly Trend PDF")
def monthly_trend_pdf(
    db: Session = Depends(get_db),
    _: User = Depends(require_role(UserRole.admin, UserRole.counselor)),
):
    students = db.query(Student).filter(
        Student.risk_score >= settings.HIGH_RISK_THRESHOLD
    ).order_by(Student.risk_score.desc()).all()

    pdf = _make_pdf_high_risk(students)
    return StreamingResponse(
        pdf,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=monthly_trend_report.pdf"}
    )
