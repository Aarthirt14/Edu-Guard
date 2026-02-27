"""
app/services/report_service.py — PDF and Excel report generation
"""
import io
from datetime import datetime
from sqlalchemy.orm import Session
from app.models.student import Student
from app.models.intervention import Intervention


def generate_high_risk_pdf(db: Session) -> bytes:
    """Generates a High Risk Students Summary PDF using reportlab."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    students = db.query(Student).filter(Student.risk_score >= 70).order_by(
        Student.risk_score.desc()
    ).all()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("EduGuard AI — High Risk Students Report", styles["Title"]))
    elements.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]))
    elements.append(Spacer(1, 0.5 * cm))

    data = [["Student ID", "Name", "Class", "Semester", "Risk Score", "Attendance%", "Marks", "Status"]]
    for s in students:
        data.append([
            s.student_code, s.name, s.student_class, s.semester,
            f"{s.risk_score:.0f}", f"{s.attendance_pct:.0f}%",
            f"{s.marks_avg:.0f}", s.intervention_status,
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0), 10),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#1F2937"), colors.HexColor("#111827")]),
        ("TEXTCOLOR",  (0, 1), (-1, -1), colors.HexColor("#E5E7EB")),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#374151")),
        ("ALIGN",      (4, 0), (-1, -1), "CENTER"),
        ("FONTSIZE",   (0, 1), (-1, -1), 8),
        ("ROWHEIGHT",  (0, 0), (-1, -1), 18),
    ]))
    elements.append(table)
    doc.build(elements)
    return buf.getvalue()


def generate_intervention_excel(db: Session) -> bytes:
    """Generates an Intervention Progress Report as XLSX."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    interventions = db.query(Intervention).order_by(Intervention.date_assigned.desc()).all()
    student_map = {s.id: s for s in db.query(Student).all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Interventions"

    headers = ["Student", "Student ID", "Type", "Assigned By", "Date", "Status", "Outcome"]
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, iv in enumerate(interventions, 2):
        student = student_map.get(iv.student_id)
        ws.cell(row=row, column=1, value=student.name if student else "Unknown")
        ws.cell(row=row, column=2, value=student.student_code if student else "")
        ws.cell(row=row, column=3, value=iv.intervention_type)
        ws.cell(row=row, column=4, value=iv.assigned_by)
        ws.cell(row=row, column=5, value=iv.date_assigned.strftime("%Y-%m-%d") if iv.date_assigned else "")
        ws.cell(row=row, column=6, value=iv.status)
        ws.cell(row=row, column=7, value=iv.outcome or "—")

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
